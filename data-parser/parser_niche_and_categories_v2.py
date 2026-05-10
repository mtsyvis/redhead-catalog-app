#!/usr/bin/env python3
"""
parser_niche_and_categories_v2.py

Batch classifier for website Niche / Categories / Language.

Key design decisions:
- Uses Anthropic Message Batches API for cheaper bulk classification.
- Default model: claude-haiku-4-5. Optional: claude-sonnet-4-6.
- Keeps Niche as a clean controlled taxonomy.
- Keeps Categories as English search tags, including site-type / attribute-like tags when useful.
- Default Excel output uses business-friendly result columns. SourceQuality can be
  good / limited / poor / protected / blocked / failed. Use --debug-output to include
  extraction/fetch/model analytics columns so bad classification can be traced back to weak input.
- Uses Selenium only as a fallback for homepage by default. Internal Selenium fallback is opt-in.
- Access-blocked pages are not sent to Claude by default. Use --classify-blocked-from-domain
  to request cautious domain-only classification.
- Use --excel-filter to enable Excel autofilter on the output worksheet.

Install dependencies:
    pip install -r data-parser/requirements.txt

Recommended Python:
    3.11

Environment:
    set ANTHROPIC_API_KEY=sk-ant-...     # Windows PowerShell: $env:ANTHROPIC_API_KEY="sk-ant-..."

Example test run:
    python parser_niche_and_categories_v2.py --input sites.xlsx --output sites_with_categories_v2.xlsx --max-sites 20

By default, output contains only processed rows. Use --output-scope all to keep the full input file in the output.

Try Sonnet for quality comparison:
    python parser_niche_and_categories_v2.py --input sites.xlsx --output sonnet_test.xlsx --max-sites 20 --model claude-sonnet-4-6

Extract only, no Claude batch:
    python parser_niche_and_categories_v2.py --input sites.xlsx --output extraction_debug.xlsx --max-sites 20 --extract-only

Default output columns:
    SourceRowNumber, Niche, Categories, Language, SourceQuality, FinalUrl, Error

Use --debug-output to include technical diagnostics such as Selenium/fetch metrics, token usage,
cost estimate, batch info, raw model output, and normalization warnings.
"""

from __future__ import annotations

import argparse
import json
import logging
import os
import random
import re
import shutil
import subprocess
import sys
import threading
import time
import traceback
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import asdict, dataclass, field
from html import unescape
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple
from urllib.parse import urljoin, urlparse, urlunparse, urldefrag

import pandas as pd
import requests
from bs4 import BeautifulSoup
from tqdm import tqdm

# Set your Anthropic API key here or via the ANTHROPIC_API_KEY environment variable
ANTHROPIC_API_KEY = None

# ===================== TAXONOMY =====================
# Niche must answer: "What is this website mainly about?"
# Categories answer: "Which English search tags help a manager find this site?"
# Attribute-like terms such as SaaS, Blog, Marketplace, Directory, B2B, Tool, App are NOT niches;
# they are allowed in Categories when useful.
NICHES_LIST: List[str] = [
    "AI",
    "Architecture",
    "Art",
    "Alcohol",
    "Adult",
    "Automotive",
    "Beauty",
    "Betting",
    "Books",
    "Business",
    "Career",
    "Cleaning",
    "Comics",
    "Construction",
    "Crypto",
    "Cybersecurity",
    "Dating",
    "Death Care",
    "Economics",
    "Education",
    "Entertainment",
    "Events",
    "Expat",
    "Family",
    "Fashion",
    "Finance",
    "Fitness",
    "Food",
    "Furniture",
    "Gambling",
    "Gaming",
    "Gardening",
    "Government",
    "Health",
    "Home Improvement",
    "Immigration",
    "Insurance",
    "Interior Design",
    "Investment",
    "Law",
    "Lifestyle",
    "Literature",
    "Logistics",
    "Marketing",
    "Media",
    "Mental Health",
    "Movies",
    "Music",
    "News",
    "Parenting",
    "Pets",
    "Photography",
    "Politics",
    "Psychology",
    "Real Estate",
    "Religion",
    "Science",
    "Shopping",
    "Software",
    "Sports",
    "Sustainability",
    "Technology",
    "Telecom",
    "Transport",
    "Travel",
    "Wedding",
    "Wellness",
    "Writing",
]

SPECIAL_NICHES = ["UNKNOWN"]
ALLOWED_NICHES = set(NICHES_LIST + SPECIAL_NICHES)

# These are allowed in Categories, not in Niche. The model can use them if they help search/filtering.
CATEGORY_ATTRIBUTE_HINTS: List[str] = [
    "App",
    "B2B",
    "B2C",
    "Blog",
    "Booking platform",
    "Community",
    "CRM",
    "Directory",
    "Ecommerce",
    "Forum",
    "Job board",
    "Marketplace",
    "Media site",
    "Metasearch",
    "News site",
    "Platform",
    "Review site",
    "SaaS",
    "Search engine",
    "Service provider",
    "Social media",
    "Streaming",
    "Subscription box",
    "Templates",
    "Tool",
    "VPN",
    "Website builder",
]

# Very weak tags that should not be kept unless embedded in a meaningful phrase.
TOO_GENERIC_CATEGORY_TAGS = {
    "best",
    "top",
    "official",
    "website",
    "homepage",
    "online",
    "company",
    "services",
    "solutions",
    "blog",
    "news",
}

LANGUAGE_UNKNOWN = "UNKNOWN"
LANGUAGE_MULTI = "MULTI"

MODEL_CHOICES = ("claude-haiku-4-5", "claude-sonnet-4-6")
DEFAULT_MODEL = "claude-haiku-4-5"
# Estimated Anthropic Message Batch prices in USD per 1M tokens.
# Used only for local run cost estimation. Anthropic Console remains the billing source of truth.
BATCH_PRICING_USD_PER_MTOK = {
    "claude-haiku-4-5": {
        "input": 0.50,
        "output": 2.50,
    },
    "claude-sonnet-4-6": {
        "input": 1.50,
        "output": 7.50,
    },
}
DEFAULT_INPUT = "sites.xlsx"
DEFAULT_OUTPUT = "sites_with_categories_v2.xlsx"
DEFAULT_URL_COLUMN = None
DEFAULT_CHECKPOINT = "parser_v2_checkpoint.json"
DEFAULT_MAX_SITES = 20
DEFAULT_START_ROW = 0
DEFAULT_CONCURRENCY = 6
DEFAULT_REQUEST_TIMEOUT = 12
DEFAULT_INTERNAL_PAGES = 3
DEFAULT_MIN_TEXT_CHARS = 180
DEFAULT_MIN_INTERNAL_TEXT_CHARS = 120
DEFAULT_SELENIUM_TIMEOUT = 12
DEFAULT_SELENIUM_HARD_TIMEOUT = 25
DEFAULT_SELENIUM_INTERNAL_LIMIT = 2
DEFAULT_MIN_CATEGORIES = 10
DEFAULT_MAX_CATEGORIES = 20
DEFAULT_MAX_TOKENS = 900
DEFAULT_MAX_PROMPT_CHARS = 20_000
DEFAULT_POLL_INTERVAL = 30
DEFAULT_LOG_FILE = "parser_v2.log"
DEFAULT_PROXY: List[str] = []

_UC_CHROME_DEL_PATCHED = False


# ===================== DATA STRUCTURES =====================
@dataclass
class PageExtract:
    url: str
    fetch_method: str
    http_status: Optional[int]
    final_url: str
    redirected: bool
    title: str = ""
    meta_description: str = ""
    meta_keywords: str = ""
    html_lang: str = ""
    h1: List[str] = field(default_factory=list)
    h2: List[str] = field(default_factory=list)
    nav_items: List[str] = field(default_factory=list)
    breadcrumbs: List[str] = field(default_factory=list)
    paragraphs: List[str] = field(default_factory=list)
    text_length: int = 0
    error: str = ""


@dataclass
class SiteExtraction:
    row_index: int
    original_url: str
    normalized_start_url: str
    homepage: Optional[PageExtract] = None
    internal_pages: List[PageExtract] = field(default_factory=list)
    internal_pages_requested: int = 0
    internal_pages_http_succeeded: int = 0
    internal_pages_selenium_tried: int = 0
    internal_pages_selenium_succeeded: int = 0
    processing_time_ms: int = 0
    source_quality: str = "failed"  # good / limited / poor / protected / blocked / failed
    prompt_chars: int = 0
    estimated_input_tokens: int = 0
    exact_input_tokens: int = 0
    token_count_method: str = ""  # estimated / exact / skipped / failed
    error: str = ""

    @property
    def total_text_length(self) -> int:
        pages = ([self.homepage] if self.homepage else []) + self.internal_pages
        return sum(
            p.text_length
            for p in pages
            if p
            and not any(
                marker in [part.strip() for part in (p.error or "").split(";")]
                for marker in {"bot_protection_challenge", "access_blocked_page"}
            )
        )

    @property
    def internal_pages_succeeded(self) -> int:
        return len([p for p in self.internal_pages if not p.error and p.text_length > 0])


@dataclass
class ClassificationResult:
    row_index: int
    original_url: str
    custom_id: str = ""
    niche: List[str] = field(default_factory=lambda: ["UNKNOWN"])
    categories: List[str] = field(default_factory=list)
    language: str = LANGUAGE_UNKNOWN
    raw_niche: List[str] = field(default_factory=list)
    raw_categories: List[str] = field(default_factory=list)
    normalization_warnings: List[str] = field(default_factory=list)
    model_name: str = ""
    input_tokens: int = 0
    output_tokens: int = 0
    estimated_cost_usd: float = 0.0
    batch_id: str = ""
    batch_status: str = ""
    model_error: str = ""
    raw_model_text: str = ""


@dataclass
class SeleniumAvailability:
    enabled: bool
    available: bool
    error: str = ""
    user_message: str = ""


# ===================== LOGGING / CHECKPOINT =====================
def setup_logging(log_file: str) -> None:
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
        handlers=[logging.FileHandler(log_file, encoding="utf-8"), logging.StreamHandler(sys.stdout)],
    )


def load_checkpoint(path: str) -> Dict[str, Any]:
    if not path or not os.path.exists(path):
        return {}
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def save_checkpoint(path: str, payload: Dict[str, Any]) -> None:
    tmp = path + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    os.replace(tmp, path)


# ===================== URL / HTTP HELPERS =====================
def normalize_input_url(value: Any) -> str:
    s = str(value or "").strip()
    if not s or s.lower() in {"nan", "none", "null"}:
        return ""
    s = s.replace(" ", "")
    if not re.match(r"^https?://", s, flags=re.I):
        # Prefer HTTPS first. HTTP is used as fallback if HTTPS fails.
        s = "https://" + s
    return s


def candidate_start_urls(value: Any) -> List[str]:
    normalized = normalize_input_url(value)
    if not normalized:
        return []
    parsed = urlparse(normalized)
    if not parsed.netloc:
        return [normalized]

    https_url = urlunparse(parsed._replace(scheme="https"))
    http_url = urlunparse(parsed._replace(scheme="http"))

    result = []
    seen = set()
    for candidate in [https_url, http_url]:
        if candidate not in seen:
            result.append(candidate)
            seen.add(candidate)
    return result


def same_site(url_a: str, url_b: str) -> bool:
    host_a = (urlparse(url_a).netloc or "").lower().removeprefix("www.")
    host_b = (urlparse(url_b).netloc or "").lower().removeprefix("www.")
    return bool(host_a and host_a == host_b)


def clean_url(url: str) -> str:
    url, _frag = urldefrag(url)
    parsed = urlparse(url)
    # Remove common tracking query params by dropping query entirely for crawling stability.
    return urlunparse(parsed._replace(query=""))


def looks_like_html_url(url: str) -> bool:
    path = urlparse(url).path.lower()
    bad_ext = (
        ".jpg",
        ".jpeg",
        ".png",
        ".gif",
        ".webp",
        ".svg",
        ".pdf",
        ".doc",
        ".docx",
        ".xls",
        ".xlsx",
        ".zip",
        ".rar",
        ".mp4",
        ".mp3",
        ".avi",
        ".mov",
        ".css",
        ".js",
        ".xml",
        ".json",
    )
    return not path.endswith(bad_ext)


def is_bad_internal_link(url: str, anchor: str) -> bool:
    url_l = url.lower()
    anchor_l = (anchor or "").strip().lower()
    bad_parts = [
        "mailto:",
        "tel:",
        "javascript:",
        "/login",
        "/signin",
        "/sign-in",
        "/signup",
        "/register",
        "/cart",
        "/checkout",
        "/privacy",
        "/terms",
        "/cookie",
        "/contact",
        "/advertise",
        "/wp-login",
        "/author/",
        "/tag/",
        "?s=",
        "?search=",
        "/search",
        "/feed",
        "/rss",
    ]
    if any(part in url_l for part in bad_parts):
        return True
    if anchor_l in {"", "home", "homepage", "read more", "learn more", "click here", "more"}:
        return True
    return False


def http_fetch(url: str, timeout: int, proxies: Optional[Dict[str, str]] = None) -> Tuple[Optional[str], Optional[int], str, str]:
    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0 Safari/537.36"
        ),
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.8",
    }
    try:
        resp = requests.get(url, headers=headers, timeout=timeout, proxies=proxies, allow_redirects=True)
        ctype = resp.headers.get("content-type", "").lower()
        text = resp.text or ""
        if resp.status_code >= 400:
            return None, resp.status_code, resp.url, f"http_status_{resp.status_code}"
        if "text/html" not in ctype and "application/xhtml" not in ctype and len(text) < 500:
            return None, resp.status_code, resp.url, f"non_html_content_type:{ctype}"
        if len(text) < 200:
            return None, resp.status_code, resp.url, "too_short_html"
        return text, resp.status_code, resp.url, ""
    except Exception as e:
        return None, None, url, f"http_exception:{type(e).__name__}:{e}"


# ===================== SELENIUM FALLBACK =====================
def patch_undetected_chromedriver_cleanup(uc: Any) -> None:
    global _UC_CHROME_DEL_PATCHED
    if _UC_CHROME_DEL_PATCHED:
        return

    original_del = getattr(uc.Chrome, "__del__", None)

    def safe_del(self):
        try:
            if original_del is not None:
                original_del(self)
            else:
                try:
                    self.quit()
                except Exception:
                    pass
        except OSError as e:
            if getattr(e, "winerror", None) == 6:
                logging.debug("Ignoring known undetected_chromedriver WinError 6 during Chrome.__del__ cleanup")
                return
            logging.debug("Ignoring undetected_chromedriver Chrome.__del__ OSError during cleanup: %s", e)
        except Exception as e:
            logging.debug("Ignoring undetected_chromedriver Chrome.__del__ error during cleanup: %s", e)

    def safe_ensure_close(cls, self):
        try:
            if (
                hasattr(self, "service")
                and hasattr(self.service, "process")
                and hasattr(self.service.process, "kill")
            ):
                self.service.process.kill()
        except OSError as e:
            if getattr(e, "winerror", None) == 6:
                logging.debug("Ignoring known undetected_chromedriver WinError 6 during Chrome._ensure_close cleanup")
                return
            logging.debug("Ignoring undetected_chromedriver Chrome._ensure_close OSError during cleanup: %s", e)
        except Exception as e:
            logging.debug("Ignoring undetected_chromedriver Chrome._ensure_close error during cleanup: %s", e)

    uc.Chrome.__del__ = safe_del
    if hasattr(uc.Chrome, "_ensure_close"):
        uc.Chrome._ensure_close = classmethod(safe_ensure_close)
    _UC_CHROME_DEL_PATCHED = True


def safe_quit_driver(driver: Any) -> None:
    if driver is None:
        return
    try:
        driver.quit()
    except OSError as e:
        if getattr(e, "winerror", None) == 6:
            logging.debug("Ignoring known Selenium WinError 6 during driver.quit()")
            return
        logging.debug("Ignoring Selenium driver quit OSError: %s", e)
    except Exception as e:
        logging.debug("Ignoring Selenium driver quit error: %s", e)


def stop_page_load(driver: Any) -> None:
    try:
        driver.execute_script("window.stop();")
    except Exception as e:
        logging.debug("Ignoring Selenium window.stop() error: %s", e)


def get_selenium_page_source(driver: Any) -> str:
    try:
        source = driver.page_source or ""
        return source
    except Exception as e:
        logging.debug("Could not read Selenium page_source: %s", e)
        return ""


def build_chrome_options(uc: Any, proxy: Optional[str] = None) -> Any:
    options = uc.ChromeOptions()
    options.page_load_strategy = "eager"
    options.add_argument("--headless=new")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-gpu")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument("--no-first-run")
    options.add_argument("--no-default-browser-check")
    options.add_argument("--disable-extensions")
    options.add_argument("--disable-background-networking")
    options.add_argument("--disable-sync")
    options.add_argument("--disable-default-apps")
    if proxy:
        options.add_argument(f"--proxy-server={proxy}")
    return options


def create_uc_driver(
    uc: Any,
    proxy: Optional[str] = None,
    chrome_version_main: Optional[int] = None,
) -> Any:
    kwargs = {
        "options": build_chrome_options(uc, proxy),
        "headless": True,
        "use_subprocess": True,
    }
    if chrome_version_main is not None:
        kwargs["version_main"] = chrome_version_main
    return uc.Chrome(**kwargs)


def check_selenium_availability(mode: str) -> SeleniumAvailability:
    if mode == "off":
        return SeleniumAvailability(
            enabled=False,
            available=False,
            user_message="Selenium disabled by --selenium-mode off",
        )

    try:
        import undetected_chromedriver as uc

        patch_undetected_chromedriver_cleanup(uc)
        from selenium.webdriver.common.by import By  # noqa: F401
        from selenium.webdriver.support import expected_conditions as EC  # noqa: F401
        from selenium.webdriver.support.ui import WebDriverWait  # noqa: F401
    except ModuleNotFoundError as e:
        if getattr(e, "name", "") == "distutils":
            message = (
                "Selenium fallback is unavailable because Python 3.12 removed distutils. "
                "Recommended: use Python 3.11 for this parser, or run "
                "`python -m pip install --upgrade setuptools wheel` and retry."
            )
        else:
            message = (
                f"Selenium fallback is unavailable: {type(e).__name__}: {e}. "
                "Install parser dependencies from data-parser/requirements.txt."
            )
        if mode == "required":
            raise RuntimeError(message)
        return SeleniumAvailability(enabled=False, available=False, error=f"{type(e).__name__}:{e}", user_message=message)
    except Exception as e:
        message = (
            f"Selenium fallback is unavailable: {type(e).__name__}: {e}. "
            "Install parser dependencies from data-parser/requirements.txt."
        )
        if mode == "required":
            raise RuntimeError(message)
        return SeleniumAvailability(enabled=False, available=False, error=f"{type(e).__name__}:{e}", user_message=message)

    return SeleniumAvailability(enabled=True, available=True)


def _existing_executable_path(path: Optional[str]) -> Optional[str]:
    if not path:
        return None
    candidate = os.path.expandvars(str(path).strip().strip('"'))
    if candidate and os.path.isfile(candidate):
        return candidate
    return None


def find_chrome_executable() -> Optional[str]:
    if sys.platform.startswith("win"):
        try:
            import winreg

            registry_keys = [
                (winreg.HKEY_CURRENT_USER, r"Software\Microsoft\Windows\CurrentVersion\App Paths\chrome.exe"),
                (winreg.HKEY_LOCAL_MACHINE, r"Software\Microsoft\Windows\CurrentVersion\App Paths\chrome.exe"),
                (winreg.HKEY_LOCAL_MACHINE, r"Software\WOW6432Node\Microsoft\Windows\CurrentVersion\App Paths\chrome.exe"),
            ]
            for hive, key_path in registry_keys:
                try:
                    with winreg.OpenKey(hive, key_path) as key:
                        value, _value_type = winreg.QueryValueEx(key, "")
                    existing = _existing_executable_path(value)
                    if existing:
                        return existing
                except OSError as e:
                    logging.debug("Chrome registry lookup failed for %s: %s", key_path, e)
        except Exception as e:
            logging.debug("Chrome registry lookup unavailable: %s", e)

    for env_name in ["ProgramFiles", "ProgramFiles(x86)", "LocalAppData"]:
        base = os.environ.get(env_name)
        if base:
            existing = _existing_executable_path(os.path.join(base, "Google", "Chrome", "Application", "chrome.exe"))
            if existing:
                return existing

    for command in ["chrome", "chrome.exe"]:
        existing = _existing_executable_path(shutil.which(command))
        if existing:
            return existing

    return None


def get_windows_file_product_version(path: str) -> Optional[str]:
    if not sys.platform.startswith("win"):
        return None
    try:
        escaped_path = path.replace("'", "''")
        completed = subprocess.run(
            ["powershell", "-NoProfile", "-Command", f"(Get-Item -LiteralPath '{escaped_path}').VersionInfo.ProductVersion"],
            check=False,
            capture_output=True,
            text=True,
            timeout=5,
        )
        output = completed.stdout.strip()
        if output:
            return output
    except Exception as e:
        logging.debug("Chrome ProductVersion lookup failed for %s: %s", path, e)
    return None


def infer_chrome_version_from_install_dir(chrome_exe_path: str) -> Optional[str]:
    try:
        install_dir = Path(chrome_exe_path).parent
        version_dirs: List[Tuple[Tuple[int, int, int, int], str]] = []
        for child in install_dir.iterdir():
            if not child.is_dir():
                continue
            if not re.match(r"^\d+\.\d+\.\d+\.\d+$", child.name):
                continue
            try:
                version_key = tuple(int(part) for part in child.name.split("."))
            except ValueError:
                continue
            if len(version_key) == 4:
                version_dirs.append((version_key, child.name))
        if version_dirs:
            return max(version_dirs, key=lambda item: item[0])[1]
    except Exception as e:
        logging.debug("Chrome install directory version inference failed for %s: %s", chrome_exe_path, e)
    return None


def parse_major_version(text: str) -> Optional[int]:
    match = re.search(r"(\d+)\.", text or "")
    return int(match.group(1)) if match else None


def _run_chrome_version_command(candidate: str) -> Optional[int]:
    try:
        completed = subprocess.run(
            [candidate, "--version"],
            check=False,
            capture_output=True,
            text=True,
            timeout=5,
        )
        return parse_major_version(f"{completed.stdout} {completed.stderr}")
    except Exception as e:
        logging.debug("Chrome version command failed for %s: %s", candidate, e)
    return None


def _detect_chrome_version_main(chrome_path: Optional[str]) -> Optional[int]:
    candidates = []
    if chrome_path:
        candidates.append(chrome_path)

    for candidate in candidates:
        product_version = get_windows_file_product_version(candidate)
        major = parse_major_version(product_version or "")
        if major:
            return major

        inferred_version = infer_chrome_version_from_install_dir(candidate)
        major = parse_major_version(inferred_version or "")
        if major:
            return major

        major = _run_chrome_version_command(candidate)
        if major:
            return major

    for candidate in ["chrome", "chrome.exe", "google-chrome"]:
        major = _run_chrome_version_command(candidate)
        if major:
            return major

    return None


def detect_chrome_version_main() -> Optional[int]:
    return _detect_chrome_version_main(find_chrome_executable())


def resolve_chrome_version_main(cli_value: Optional[int]) -> Optional[int]:
    if cli_value:
        logging.info("Using Chrome major version from --chrome-version-main: %s", cli_value)
        return cli_value

    chrome_path = find_chrome_executable()
    if chrome_path:
        logging.info("Detected Chrome executable: %s", chrome_path)

    detected = _detect_chrome_version_main(chrome_path)
    if detected:
        logging.info("Detected Chrome major version: %s", detected)
        return detected

    logging.warning(
        "Chrome major version was not detected; undetected_chromedriver will choose automatically. "
        "If Selenium version mismatch happens, pass --chrome-version-main manually."
    )
    return None


def parse_current_browser_major_from_error(message: str) -> Optional[int]:
    match = re.search(r"Current browser version is\s+(\d+)\.", message)
    if match:
        return int(match.group(1))
    return None


def format_selenium_error(error: Exception) -> str:
    message = str(error)
    is_version_mismatch = (
        "This version of ChromeDriver only supports Chrome version" in message
        or "Current browser version is" in message
    )
    if is_version_mismatch:
        current_major = parse_current_browser_major_from_error(message)
        if current_major:
            return (
                "selenium_version_mismatch:ChromeDriver/Chrome version mismatch; "
                f"try --chrome-version-main {current_major}"
            )
        return (
            "selenium_version_mismatch:ChromeDriver/Chrome version mismatch; "
            "try --chrome-version-main <major_version>"
        )
    return f"selenium_exception:{type(error).__name__}:{error}"


def selenium_fetch_once(
    url: str,
    timeout: int,
    proxy: Optional[str] = None,
    chrome_version_main: Optional[int] = None,
) -> Tuple[Optional[str], str, str]:
    try:
        import undetected_chromedriver as uc

        patch_undetected_chromedriver_cleanup(uc)
        from selenium.common.exceptions import TimeoutException
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.webdriver.support.ui import WebDriverWait
    except ModuleNotFoundError as e:
        if getattr(e, "name", "") == "distutils":
            return None, url, (
                "selenium_import_failed:distutils_missing:"
                "Python 3.12 removed distutils. Use Python 3.11 or install setuptools/wheel."
            )
        return None, url, f"selenium_import_failed:{type(e).__name__}:{e}"
    except Exception as e:
        return None, url, f"selenium_import_failed:{type(e).__name__}:{e}"

    driver = None
    try:
        try:
            driver = create_uc_driver(uc, proxy=proxy, chrome_version_main=chrome_version_main)
        except Exception as e:
            retry_version = parse_current_browser_major_from_error(str(e))
            if retry_version is not None and retry_version != chrome_version_main:
                logging.warning("Retrying Selenium with detected Chrome major version from error: %s", retry_version)
                driver = create_uc_driver(uc, proxy=proxy, chrome_version_main=retry_version)
            else:
                raise

        driver.set_page_load_timeout(timeout)
        page_load_timed_out = False

        try:
            driver.get(url)
        except TimeoutException:
            page_load_timed_out = True
            logging.debug(
                "Selenium page load timed out for %s after %ss; trying to use partial page_source",
                url,
                timeout,
            )
            stop_page_load(driver)

        try:
            WebDriverWait(driver, min(timeout, 5)).until(
                EC.presence_of_element_located((By.TAG_NAME, "body"))
            )
        except TimeoutException:
            logging.debug("Selenium body wait timed out for %s; trying page_source anyway", url)

        time.sleep(0.2)
        html = get_selenium_page_source(driver)
        try:
            final_url = driver.current_url or url
        except Exception as e:
            logging.debug("Could not read Selenium current_url: %s", e)
            final_url = url

        if html and len(html) >= 200:
            if page_load_timed_out:
                return html, final_url, "selenium_page_load_timeout_but_source_available"
            return html, final_url, ""

        if page_load_timed_out:
            return None, final_url, "selenium_page_load_timeout_no_usable_source"
        return None, final_url, "selenium_no_usable_source"
    except Exception as e:
        if type(e).__name__ == "TimeoutException":
            return None, url, "selenium_page_load_timeout_no_usable_source"
        return None, url, format_selenium_error(e)
    finally:
        safe_quit_driver(driver)


def selenium_fetch_with_hard_timeout(
    url: str,
    timeout: int,
    hard_timeout: int,
    proxy: Optional[str] = None,
    chrome_version_main: Optional[int] = None,
) -> Tuple[Optional[str], str, str]:
    result: Dict[str, Any] = {"html": None, "final_url": url, "error": ""}

    def run() -> None:
        html, final_url, err = selenium_fetch_once(
            url,
            timeout=timeout,
            proxy=proxy,
            chrome_version_main=chrome_version_main,
        )
        result["html"] = html
        result["final_url"] = final_url
        result["error"] = err

    thread = threading.Thread(target=run, daemon=True)
    thread.start()
    thread.join(timeout=hard_timeout)
    if thread.is_alive():
        return None, url, f"selenium_hard_timeout_{hard_timeout}s"
    return result["html"], result["final_url"], result["error"]


# ===================== HTML EXTRACTION =====================
def compact_text(text: str, max_len: int = 500) -> str:
    if not text:
        return ""
    text = unescape(text)
    text = re.sub(r"\s+", " ", text).strip()
    if len(text) > max_len:
        text = text[: max_len - 1].rstrip() + "…"
    return text


def unique_keep_order(items: Iterable[str], limit: int) -> List[str]:
    seen = set()
    result = []
    for item in items:
        cleaned = compact_text(item)
        key = cleaned.lower()
        if not cleaned or key in seen:
            continue
        seen.add(key)
        result.append(cleaned)
        if len(result) >= limit:
            break
    return result


def extract_links(soup: BeautifulSoup, base_url: str, limit: int) -> List[str]:
    if limit <= 0:
        return []

    candidates: List[Tuple[int, str, str]] = []

    def add_link(a: Any, context_score: int) -> None:
        href = a.get("href") if a else None
        if not href:
            return
        anchor = compact_text(a.get_text(" ", strip=True), 120)
        full = clean_url(urljoin(base_url, href))
        if not looks_like_html_url(full):
            return
        if not same_site(base_url, full):
            return
        if clean_url(full).rstrip("/") == clean_url(base_url).rstrip("/"):
            return
        if is_bad_internal_link(full, anchor):
            return
        url_l = full.lower()
        anchor_l = anchor.lower()
        score = context_score
        strong_terms = [
            "category",
            "categories",
            "topic",
            "topics",
            "section",
            "sections",
            "blog",
            "article",
            "articles",
            "resources",
            "services",
            "industries",
            "solutions",
            "about",
            "guide",
            "guides",
        ]
        if any(term in url_l or term in anchor_l for term in strong_terms):
            score += 20
        # Prefer meaningful anchor text over icon/header links.
        if len(anchor_l) >= 4:
            score += min(len(anchor_l), 40) // 4
        candidates.append((score, full, anchor))

    for nav in soup.find_all(["nav", "header"]):
        for a in nav.find_all("a", href=True):
            add_link(a, 30)

    for menu in soup.find_all(["ul", "ol", "div"]):
        classes = " ".join(menu.get("class") or []).lower()
        ident = (menu.get("id") or "").lower()
        if any(k in classes for k in ["menu", "nav", "category", "topic", "section"]) or any(
            k in ident for k in ["menu", "nav", "category", "topic", "section"]
        ):
            for a in menu.find_all("a", href=True):
                add_link(a, 20)

    for a in soup.find_all("a", href=True):
        add_link(a, 0)

    # Highest score first, unique URLs.
    candidates.sort(key=lambda x: x[0], reverse=True)
    result = []
    seen = set()
    for _score, url, _anchor in candidates:
        key = url.rstrip("/")
        if key in seen:
            continue
        seen.add(key)
        result.append(url)
        if len(result) >= limit:
            break
    return result


def parse_html(html: str, url: str, fetch_method: str, http_status: Optional[int], final_url: str) -> Tuple[PageExtract, List[str]]:
    soup = BeautifulSoup(html or "", "html.parser")

    for tag in soup(["script", "style", "noscript", "svg", "canvas", "iframe"]):
        tag.decompose()

    title = compact_text(soup.title.string if soup.title and soup.title.string else "", 300)

    meta_desc = ""
    meta_desc_tag = soup.find("meta", attrs={"name": re.compile(r"^description$", re.I)})
    if meta_desc_tag and meta_desc_tag.get("content"):
        meta_desc = compact_text(meta_desc_tag.get("content", ""), 500)

    meta_keywords = ""
    meta_keywords_tag = soup.find("meta", attrs={"name": re.compile(r"^keywords$", re.I)})
    if meta_keywords_tag and meta_keywords_tag.get("content"):
        meta_keywords = compact_text(meta_keywords_tag.get("content", ""), 500)

    html_lang = ""
    try:
        if soup.html and soup.html.has_attr("lang"):
            html_lang = compact_text(str(soup.html.get("lang") or ""), 40)
    except Exception:
        html_lang = ""

    h1 = unique_keep_order((h.get_text(" ", strip=True) for h in soup.find_all("h1")), limit=8)
    h2 = unique_keep_order((h.get_text(" ", strip=True) for h in soup.find_all("h2")), limit=16)

    nav_items: List[str] = []
    for nav in soup.find_all("nav"):
        nav_items.extend([a.get_text(" ", strip=True) for a in nav.find_all("a")])
        nav_items.extend([li.get_text(" ", strip=True) for li in nav.find_all("li")])
    for menu in soup.find_all(["ul", "ol", "div"]):
        classes = " ".join(menu.get("class") or []).lower()
        ident = (menu.get("id") or "").lower()
        if any(k in classes for k in ["menu", "nav", "category", "topic"]) or any(
            k in ident for k in ["menu", "nav", "category", "topic"]
        ):
            nav_items.extend([li.get_text(" ", strip=True) for li in menu.find_all("li")])
            nav_items.extend([a.get_text(" ", strip=True) for a in menu.find_all("a")])
    nav_items = unique_keep_order(nav_items, limit=40)

    breadcrumbs: List[str] = []
    for node in soup.find_all(class_=re.compile(r"breadcrumb", re.I)):
        breadcrumbs.append(node.get_text(" ", strip=True))
    breadcrumbs = unique_keep_order(breadcrumbs, limit=10)

    paragraphs = unique_keep_order((p.get_text(" ", strip=True) for p in soup.find_all("p")), limit=18)

    text_parts = [title, meta_desc, meta_keywords] + h1 + h2 + nav_items + breadcrumbs + paragraphs
    text_length = len(" ".join([p for p in text_parts if p]))

    page = PageExtract(
        url=url,
        fetch_method=fetch_method,
        http_status=http_status,
        final_url=final_url,
        redirected=clean_url(url).rstrip("/") != clean_url(final_url).rstrip("/"),
        title=title,
        meta_description=meta_desc,
        meta_keywords=meta_keywords,
        html_lang=html_lang,
        h1=h1,
        h2=h2,
        nav_items=nav_items,
        breadcrumbs=breadcrumbs,
        paragraphs=paragraphs,
        text_length=text_length,
    )

    links = extract_links(soup, final_url or url, limit=50)
    return page, links


BOT_PROTECTION_PATTERNS = [
    "verifying you are a human",
    "needs to review the security of your connection",
    "checking if the site connection is secure",
    "checking your browser",
    "just a moment",
    "please wait while we verify",
    "enable javascript and cookies",
    "cf-browser-verification",
    "cloudflare ray id",
    "attention required",
    "ddos-guard",
    "bot verification",
    "human verification",
]

BLOCKED_PAGE_PATTERNS = [
    "website blocked",
    "access to this website has been denied",
    "akses ke situs ini diblokir",
    "diblokir oleh pemerintah",
    "trustpositif",
    "trust+",
    "konten negatif",
    "negative content",
    "internet positif",
    "situs ini diblokir",
    "blocked by government",
    "blocked by your country",
    "blocked in your country",
    "this website is blocked",
    "access denied by",
    "content is not available in your country",
    "not available in your region",
]


def append_error(existing: str, new_error: str) -> str:
    existing = (existing or "").strip()
    new_error = (new_error or "").strip()
    if not existing:
        return new_error
    if not new_error:
        return existing
    parts = [p.strip() for p in existing.split(";") if p.strip()]
    if new_error in parts:
        return existing
    return existing + "; " + new_error


def combined_page_text(page: Optional[PageExtract]) -> str:
    if not page:
        return ""
    parts = [
        page.title,
        page.meta_description,
        page.meta_keywords,
        *page.h1,
        *page.h2,
        *page.nav_items,
        *page.breadcrumbs,
        *page.paragraphs,
    ]
    return " ".join(str(p) for p in parts if p)


def is_bot_protection_text(text: str) -> bool:
    normalized = re.sub(r"\s+", " ", (text or "").lower()).strip()
    if not normalized:
        return False
    return any(pattern in normalized for pattern in BOT_PROTECTION_PATTERNS)


def is_bot_protection_page(page: Optional[PageExtract]) -> bool:
    if not page:
        return False
    return is_bot_protection_text(combined_page_text(page))


def is_blocked_page_text(text: str) -> bool:
    normalized = re.sub(r"\s+", " ", (text or "").lower()).strip()
    if not normalized:
        return False
    return any(pattern in normalized for pattern in BLOCKED_PAGE_PATTERNS)


def is_blocked_page(page: Optional[PageExtract]) -> bool:
    if not page:
        return False
    return is_blocked_page_text(combined_page_text(page))


def page_has_error(page: PageExtract, error_code: str) -> bool:
    return error_code in [part.strip() for part in (page.error or "").split(";")]


def page_to_prompt_payload(page: PageExtract) -> Dict[str, Any]:
    return {
        "url": page.final_url or page.url,
        "fetch_method": page.fetch_method,
        "title": page.title,
        "meta_description": page.meta_description,
        "meta_keywords": page.meta_keywords,
        "html_lang": page.html_lang,
        "h1": page.h1,
        "h2": page.h2,
        "nav_items": page.nav_items,
        "breadcrumbs": page.breadcrumbs,
        "paragraphs": page.paragraphs,
    }


# ===================== SITE CRAWLING =====================
def choose_proxy(proxies: Sequence[str]) -> Optional[str]:
    return random.choice(list(proxies)) if proxies else None


def fetch_page(
    url: str,
    *,
    request_timeout: int,
    selenium_timeout: int,
    selenium_hard_timeout: int,
    use_selenium_fallback: bool,
    proxies: Sequence[str],
    min_text_chars: int,
    chrome_version_main: Optional[int] = None,
) -> Tuple[Optional[PageExtract], List[str], str]:
    proxy = choose_proxy(proxies)
    proxy_dict = {"http": proxy, "https": proxy} if proxy else None

    html, status, final_url, err = http_fetch(url, timeout=request_timeout, proxies=proxy_dict)
    if html:
        page, links = parse_html(html, url=url, fetch_method="http", http_status=status, final_url=final_url)
        if is_bot_protection_page(page):
            return page, links, ""
        if is_blocked_page(page):
            return page, links, ""
        if page.text_length >= min_text_chars or not use_selenium_fallback:
            return page, links, ""
        # Text exists but is too weak. Let Selenium try to render it.
        err = f"weak_http_text:{page.text_length}"

    if not use_selenium_fallback:
        return None, [], err or "http_failed"

    selenium_proxy = choose_proxy(proxies)
    html2, final_url2, selenium_err = selenium_fetch_with_hard_timeout(
        url,
        timeout=selenium_timeout,
        hard_timeout=selenium_hard_timeout,
        proxy=selenium_proxy,
        chrome_version_main=chrome_version_main,
    )
    if html2:
        page2, links2 = parse_html(html2, url=url, fetch_method="selenium", http_status=status, final_url=final_url2)
        if page2.text_length > 0:
            return page2, links2, ""
    return None, [], selenium_err or err or "fetch_failed"


def evaluate_source_quality(extraction: SiteExtraction, min_text_chars: int) -> str:
    if extraction.homepage and is_bot_protection_page(extraction.homepage):
        return "protected"
    if extraction.homepage and is_blocked_page(extraction.homepage):
        return "blocked"
    if extraction.error and not extraction.homepage:
        return "failed"
    total = extraction.total_text_length
    if total >= max(1500, min_text_chars * 4) and extraction.internal_pages_succeeded >= 2:
        return "good"
    if total >= max(700, min_text_chars * 2):
        return "limited"
    if total >= min_text_chars:
        return "poor"
    return "failed"


def crawl_site(
    row_index: int,
    original_url: str,
    args: argparse.Namespace,
    chrome_version_main: Optional[int] = None,
) -> SiteExtraction:
    start_time = time.time()
    extraction = SiteExtraction(
        row_index=row_index,
        original_url=str(original_url or "").strip(),
        normalized_start_url=normalize_input_url(original_url),
        internal_pages_requested=max(0, args.internal_pages),
    )

    candidates = candidate_start_urls(original_url)
    if not candidates:
        extraction.error = "empty_url"
        extraction.processing_time_ms = int((time.time() - start_time) * 1000)
        return extraction

    homepage = None
    homepage_links: List[str] = []
    last_error = ""
    selenium_availability: SeleniumAvailability = args.selenium_availability
    homepage_selenium_enabled = args.selenium_mode != "off" and selenium_availability.available
    for candidate in candidates:
        page, links, err = fetch_page(
            candidate,
            request_timeout=args.request_timeout,
            selenium_timeout=args.selenium_timeout,
            selenium_hard_timeout=args.selenium_hard_timeout,
            use_selenium_fallback=homepage_selenium_enabled,
            proxies=args.proxies,
            min_text_chars=args.min_text_chars,
            chrome_version_main=chrome_version_main,
        )
        if page:
            homepage = page
            homepage_links = links
            break
        if args.selenium_mode == "auto" and not selenium_availability.available:
            err = f"{err}; selenium_unavailable" if err else "selenium_unavailable"
        last_error = err

    if not homepage:
        extraction.error = last_error or "homepage_fetch_failed"
        extraction.processing_time_ms = int((time.time() - start_time) * 1000)
        extraction.source_quality = evaluate_source_quality(extraction, args.min_text_chars)
        return extraction

    extraction.homepage = homepage
    if is_bot_protection_page(homepage):
        homepage.error = append_error(homepage.error, "bot_protection_challenge")
        extraction.source_quality = "protected"
        extraction.error = append_error(extraction.error, "bot_protection_challenge")
        extraction.internal_pages = []
        extraction.processing_time_ms = int((time.time() - start_time) * 1000)
        logging.info(
            "Bot protection detected for %s final_url=%s; skipping normal content classification",
            extraction.original_url,
            homepage.final_url if homepage else "",
        )
        return extraction

    if is_blocked_page(homepage):
        homepage.error = append_error(homepage.error, "access_blocked_page")
        extraction.source_quality = "blocked"
        extraction.error = append_error(extraction.error, "access_blocked_page")
        extraction.internal_pages = []
        extraction.processing_time_ms = int((time.time() - start_time) * 1000)
        logging.info(
            "Access/block page detected for %s final_url=%s; skipping normal content classification",
            extraction.original_url,
            homepage.final_url if homepage else "",
        )
        return extraction

    selected_links = homepage_links[: max(0, args.internal_pages)]
    internal_selenium_enabled = args.selenium_internal_fallback and selenium_availability.available
    selenium_internal_remaining = max(0, args.selenium_internal_limit if internal_selenium_enabled else 0)

    for link in selected_links:
        if len(extraction.internal_pages) >= args.internal_pages:
            break
        proxy = choose_proxy(args.proxies)
        proxy_dict = {"http": proxy, "https": proxy} if proxy else None
        html, status, final_url, err = http_fetch(link, timeout=args.request_timeout, proxies=proxy_dict)
        if html:
            page, _links = parse_html(html, url=link, fetch_method="http", http_status=status, final_url=final_url)
            if is_bot_protection_page(page):
                page.error = append_error(page.error, "bot_protection_challenge")
                extraction.internal_pages.append(page)
                logging.debug("Internal bot protection detected for %s final_url=%s", link, page.final_url)
                continue
            if is_blocked_page(page):
                page.error = append_error(page.error, "access_blocked_page")
                extraction.internal_pages.append(page)
                logging.debug("Internal access/block page detected for %s final_url=%s", link, page.final_url)
                continue
            if page.text_length >= args.min_internal_text_chars:
                extraction.internal_pages.append(page)
                extraction.internal_pages_http_succeeded += 1
                continue

        if internal_selenium_enabled and selenium_internal_remaining > 0:
            extraction.internal_pages_selenium_tried += 1
            selenium_internal_remaining -= 1
            html2, final_url2, selenium_err = selenium_fetch_with_hard_timeout(
                link,
                timeout=args.selenium_timeout,
                hard_timeout=args.selenium_hard_timeout,
                proxy=choose_proxy(args.proxies),
                chrome_version_main=chrome_version_main,
            )
            if html2:
                page2, _links2 = parse_html(html2, url=link, fetch_method="selenium", http_status=status, final_url=final_url2)
                if is_bot_protection_page(page2):
                    page2.error = append_error(page2.error, "bot_protection_challenge")
                    extraction.internal_pages.append(page2)
                    logging.debug("Internal bot protection detected for %s final_url=%s", link, page2.final_url)
                    continue
                if is_blocked_page(page2):
                    page2.error = append_error(page2.error, "access_blocked_page")
                    extraction.internal_pages.append(page2)
                    logging.debug("Internal access/block page detected for %s final_url=%s", link, page2.final_url)
                    continue
                if page2.text_length >= args.min_internal_text_chars:
                    extraction.internal_pages.append(page2)
                    extraction.internal_pages_selenium_succeeded += 1
                    continue
            logging.debug("Internal Selenium failed for %s: %s", link, selenium_err)
        else:
            logging.debug("Internal HTTP failed/weak for %s: %s", link, err)

    extraction.processing_time_ms = int((time.time() - start_time) * 1000)
    extraction.source_quality = evaluate_source_quality(extraction, args.min_text_chars)
    return extraction


# ===================== CLAUDE PROMPT / VALIDATION =====================
def build_system_prompt() -> str:
    return (
        "You are a strict website classifier for an advertising placement catalog. "
        "Return only one valid JSON object. Do not include markdown, comments, explanations, or extra text. "
        "Use English for Niche and Categories, even when the website content is not English. "
        "Language must be the primary content language as an ISO 639-1 uppercase code, or MULTI for genuinely multilingual sites, or UNKNOWN if unclear. "
        "Follow the user's schema and constraints exactly."
    )


def build_user_prompt(
    extraction: SiteExtraction,
    max_prompt_chars: int,
    min_categories: int,
    max_categories: int,
) -> str:
    payload = {
        "original_url": extraction.original_url,
        "source_quality": extraction.source_quality,
        "niche_whitelist": NICHES_LIST,
        "special_niches": SPECIAL_NICHES,
        "category_attribute_hints": CATEGORY_ATTRIBUTE_HINTS,
        "rules": {
          "output": [
            "Return only one valid JSON object.",
            "No markdown, no explanation, no extra text.",
            "Use exactly these keys: Niche, Categories, Language."
          ],
          "niche": [
            "Niche must contain 1 to 3 strings.",
            "Every Niche value must be copied exactly from niche_whitelist or special_niches.",
            "Never invent niche values, sub-niches, variants, or more specific labels.",
            "If a specific topic is not in niche_whitelist, choose the closest broader exact whitelist value.",
            "If no whitelist value reasonably matches, use UNKNOWN."
          ],
          "categories": [
            f"Target {min_categories} to {max_categories} English search tags when there is enough distinct content.",
            "Return fewer categories if needed; never add filler.",
            "Categories should help managers search candidate websites for client ad placement requests.",
            "Use specific searchable phrases: subniches, synonyms, services, products, audience terms, site format attributes.",
            "Do not include language codes or language names.",
            "Avoid standalone generic tags unless part of a meaningful phrase."
          ],
          "language": [
            "Language must be the main content language as an ISO 639-1 uppercase code, e.g. EN, DE, FR.",
            "Use UNKNOWN if the language cannot be determined."
          ]
        },
        "required_json_shape": {
            "Niche": ["<exact whitelist niche>"],
            "Categories": [
                "B2B SaaS",
                "accounting software",
                "invoice management",
                "tax reporting",
                "business finance",
                "expense tracking",
                "financial automation",
                "small business accounting",
            ],
            "Language": "EN",
        },
        "extracted_pages": [],
    }
    if extraction.homepage:
        payload["extracted_pages"].append({"type": "homepage", **page_to_prompt_payload(extraction.homepage)})
    for page in extraction.internal_pages:
        if page_has_error(page, "bot_protection_challenge"):
            continue
        if page_has_error(page, "access_blocked_page"):
            continue
        payload["extracted_pages"].append({"type": "internal", **page_to_prompt_payload(page)})

    # Keep cost predictable. We include pages in priority order: homepage first, then selected internals.
    # If the prompt is too large, drop internal pages from the end before submitting.
    rendered = json.dumps(payload, ensure_ascii=False, indent=2)
    if max_prompt_chars and len(rendered) > max_prompt_chars and len(payload["extracted_pages"]) > 1:
        while len(payload["extracted_pages"]) > 1:
            payload["extracted_pages"].pop()
            rendered = json.dumps(payload, ensure_ascii=False, indent=2)
            if len(rendered) <= max_prompt_chars:
                break

    # Last-resort trim for pathological homepages. Do not trim metadata/rules; only reduce text-heavy lists.
    if max_prompt_chars and len(rendered) > max_prompt_chars and payload["extracted_pages"]:
        page = payload["extracted_pages"][0]
        page["paragraphs"] = page.get("paragraphs", [])[:6]
        page["nav_items"] = page.get("nav_items", [])[:25]
        page["h2"] = page.get("h2", [])[:8]
        rendered = json.dumps(payload, ensure_ascii=False, indent=2)

    return rendered

def build_domain_only_prompt(
    extraction: SiteExtraction,
    min_categories: int,
    max_categories: int,
) -> str:
    domain_only_max_categories = min(5, max_categories)

    payload = {
        "original_url": extraction.original_url,
        "final_url": extraction.homepage.final_url if extraction.homepage else "",
        "normalized_start_url": extraction.normalized_start_url,
        "source_quality": extraction.source_quality,
        "niche_whitelist": NICHES_LIST,
        "special_niches": SPECIAL_NICHES,
        "rules": {
            "output": [
                "Return only one valid JSON object.",
                "Use exactly these keys: Niche, Categories, Language.",
                "Do not include markdown, comments, explanations, or extra text.",
            ],
            "domain_only_source": [
                "Classify only from original_url, final_url, and normalized_start_url.",
                "Treat this as weak domain-only inference.",
                "Prefer UNKNOWN over guessing.",
                "If the domain/final URL is ambiguous, return Niche [\"UNKNOWN\"], Categories [], Language \"UNKNOWN\".",
            ],
            "niche": [
                "Niche must contain 1 to 3 strings.",
                "Every Niche value must exactly match niche_whitelist or special_niches.",
                "Do not invent Niche values, sub-niches, variants, or more specific labels.",
                "If no exact whitelist value strongly matches the domain/final URL, use [\"UNKNOWN\"].",
            ],
            "categories": [
                "Return 0 to 3 Categories by default.",
                "Use up to 5 only when the domain/final URL is very specific.",
                "Categories must be derived only from clear domain/final URL signals.",
                "Do not invent detailed Categories.",
                "Do not add generic filler tags.",
            ],
            "language": [
                "Use UNKNOWN unless the domain/final URL strongly and unambiguously indicates the primary content language.",
                "Do not infer language from TLD alone unless it is very strong and unambiguous.",
            ],
        },
        "required_json_shape": {
            "Niche": ["<exact niche from niche_whitelist or UNKNOWN>"],
            "Categories": ["<cautious English domain-derived tag>"],
            "Language": "UNKNOWN",
        },
    }

    return json.dumps(payload, ensure_ascii=False, indent=2)


def safe_load_json_object(text: str) -> Optional[Dict[str, Any]]:
    if not text:
        return None
    cleaned = text.strip()
    cleaned = re.sub(r"^```(?:json)?\s*", "", cleaned, flags=re.I)
    cleaned = re.sub(r"\s*```$", "", cleaned)
    try:
        obj = json.loads(cleaned)
        return obj if isinstance(obj, dict) else None
    except Exception:
        pass

    # Last resort: extract first JSON object.
    match = re.search(r"\{.*\}", cleaned, flags=re.S)
    if match:
        try:
            obj = json.loads(match.group(0))
            return obj if isinstance(obj, dict) else None
        except Exception:
            return None
    return None


def coerce_list(value: Any) -> List[str]:
    if value is None:
        return []
    if isinstance(value, list):
        raw = value
    elif isinstance(value, str):
        raw = re.split(r"[,;|]", value)
    else:
        raw = [str(value)]
    return [compact_text(str(x), 120) for x in raw if compact_text(str(x), 120)]


def normalize_niche_key(value: Any) -> str:
    text = str(value or "").strip()
    text = re.sub(r"[_-]+", " ", text)
    text = re.sub(r"\s+", " ", text)
    return text.lower()


def normalize_niches(value: Any) -> Tuple[List[str], List[str]]:
    raw = coerce_list(value)
    result: List[str] = []
    invalid: List[str] = []
    # Case-insensitive exact matching back to canonical whitelist value after safe whitespace normalization only.
    canonical = {normalize_niche_key(n): n for n in ALLOWED_NICHES}
    for item in raw:
        key = normalize_niche_key(item)
        if key in canonical and canonical[key] not in result:
            result.append(canonical[key])
        elif item not in invalid:
            invalid.append(item)
    result = [n for n in result if n != "UNKNOWN"][:3]
    return result or ["UNKNOWN"], invalid


def normalize_categories(value: Any, niches: List[str], max_categories: int = DEFAULT_MAX_CATEGORIES) -> List[str]:
    raw = coerce_list(value)
    result: List[str] = []
    seen = set()
    for item in raw:
        cleaned = compact_text(item, 100).strip(" ,.;:-")
        if not cleaned:
            continue
        if re.fullmatch(r"[A-Z]{2}", cleaned.upper()):
            continue
        if cleaned.lower() in {"english", "german", "french", "spanish", "russian", "indonesian", "multi", "multilingual"}:
            continue
        if cleaned.lower() in TOO_GENERIC_CATEGORY_TAGS:
            continue
        key = cleaned.lower()
        if key in seen:
            continue
        seen.add(key)
        result.append(cleaned)
        if len(result) >= max_categories:
            break
    return result


def normalize_language(value: Any) -> str:
    s = compact_text(str(value or ""), 40).upper().replace("-", "_")
    if not s or s in {"N/A", "NA", "NONE", "NULL", "UNKNOWN"}:
        return LANGUAGE_UNKNOWN
    if s in {"MULTI", "MULTILINGUAL", "MIXED"}:
        return LANGUAGE_MULTI
    # Convert EN_US / PT_BR to EN / PT.
    if re.fullmatch(r"[A-Z]{2}_[A-Z]{2}", s):
        return s[:2]
    if re.fullmatch(r"[A-Z]{2}", s):
        return s
    # Basic language names fallback.
    name_map = {
        "ENGLISH": "EN",
        "GERMAN": "DE",
        "FRENCH": "FR",
        "SPANISH": "ES",
        "ITALIAN": "IT",
        "PORTUGUESE": "PT",
        "POLISH": "PL",
        "RUSSIAN": "RU",
        "UKRAINIAN": "UK",
        "INDONESIAN": "ID",
        "DUTCH": "NL",
        "TURKISH": "TR",
    }
    return name_map.get(s, LANGUAGE_UNKNOWN)


def parse_and_validate_model_result(
    row_index: int,
    original_url: str,
    custom_id: str,
    model_name: str,
    batch_id: str,
    batch_status: str,
    raw_text: str,
    max_categories: int = DEFAULT_MAX_CATEGORIES,
) -> ClassificationResult:
    result = ClassificationResult(
        row_index=row_index,
        original_url=original_url,
        custom_id=custom_id,
        model_name=model_name,
        batch_id=batch_id,
        batch_status=batch_status,
        raw_model_text=raw_text or "",
    )
    obj = safe_load_json_object(raw_text)
    if not obj:
        result.niche = ["UNKNOWN"]
        result.categories = []
        result.language = LANGUAGE_UNKNOWN
        result.model_error = "model_returned_invalid_json"
        return result

    raw_niche_value = obj.get("Niche") or obj.get("niche") or obj.get("niches")
    raw_categories_value = obj.get("Categories") or obj.get("categories") or obj.get("Category")
    result.raw_niche = coerce_list(raw_niche_value)
    result.raw_categories = coerce_list(raw_categories_value)

    invalid_niche_values: List[str]
    result.niche, invalid_niche_values = normalize_niches(result.raw_niche)
    result.categories = normalize_categories(
        result.raw_categories,
        result.niche,
        max_categories=max_categories,
    )
    result.language = normalize_language(obj.get("Language") or obj.get("language"))
    if invalid_niche_values:
        result.normalization_warnings.append("invalid_niche_values: " + ", ".join(invalid_niche_values))
    if result.niche == ["UNKNOWN"] and result.raw_categories:
        result.normalization_warnings.append("niche_unknown_but_categories_present")
    if result.niche == ["UNKNOWN"] and result.categories:
        result.normalization_warnings.append("categories_kept_despite_unknown_niche")
    return result


# ===================== ANTHROPIC BATCH =====================
def resolve_anthropic_api_key(cli_api_key: Optional[str] = None) -> str:
    for candidate in [cli_api_key, ANTHROPIC_API_KEY, os.getenv("ANTHROPIC_API_KEY")]:
        key = (candidate or "").strip()
        if key:
            return key
    raise RuntimeError(
        "Anthropic API key is missing. Provide it via --api-key, "
        "ANTHROPIC_API_KEY constant, or ANTHROPIC_API_KEY env var."
    )


def create_anthropic_client(api_key: Optional[str] = None) -> Any:
    try:
        import anthropic
    except Exception as e:
        raise RuntimeError(
            "Anthropic SDK is not installed. Run: pip install anthropic"
        ) from e
    return anthropic.Anthropic(api_key=resolve_anthropic_api_key(api_key))


def estimate_tokens_from_text(text: str) -> int:
    # Conservative rough estimate for English-ish JSON prompts. Exact counting is available via --count-tokens.
    return max(1, int(len(text) / 4) + 1)


def calculate_batch_cost_usd(model: str, input_tokens: int, output_tokens: int) -> float:
    prices = BATCH_PRICING_USD_PER_MTOK.get(model) or BATCH_PRICING_USD_PER_MTOK.get(DEFAULT_MODEL)
    if not prices:
        return 0.0

    input_price = float(prices.get("input", 0.0))
    output_price = float(prices.get("output", 0.0))

    return round(
        (max(0, input_tokens) / 1_000_000) * input_price
        + (max(0, output_tokens) / 1_000_000) * output_price,
        8,
    )


def count_input_tokens_exact(client: Any, model: str, system_prompt: str, user_prompt: str) -> Tuple[int, str]:
    try:
        response = client.messages.count_tokens(
            model=model,
            system=system_prompt,
            messages=[{"role": "user", "content": user_prompt}],
        )
        return int(getattr(response, "input_tokens", 0) or 0), "exact"
    except Exception as e:
        logging.warning("Token counting failed: %s", e)
        return 0, "failed"


def should_send_extraction_to_model(
    extraction: SiteExtraction,
    classify_protected_from_domain: bool,
    classify_blocked_from_domain: bool,
) -> bool:
    if not extraction.homepage:
        return False
    if extraction.source_quality == "failed":
        return False
    if extraction.source_quality == "protected":
        return bool(classify_protected_from_domain)
    if extraction.source_quality == "blocked":
        return bool(classify_blocked_from_domain)
    return True


def create_batch_requests(
    extractions: List[SiteExtraction],
    model: str,
    max_tokens: int,
    max_prompt_chars: int,
    min_categories: int,
    max_categories: int,
    client: Optional[Any] = None,
    count_tokens: bool = False,
    classify_protected_from_domain: bool = False,
    classify_blocked_from_domain: bool = False,
) -> Tuple[List[Any], Dict[str, SiteExtraction]]:
    try:
        from anthropic.types.message_create_params import MessageCreateParamsNonStreaming
        from anthropic.types.messages.batch_create_params import Request
    except Exception as e:
        raise RuntimeError(
            "Anthropic SDK batch types are unavailable. Upgrade SDK: pip install --upgrade anthropic"
        ) from e

    requests_payload = []
    mapping: Dict[str, SiteExtraction] = {}
    system_prompt = build_system_prompt()

    for seq, extraction in enumerate(extractions, start=1):
        if not should_send_extraction_to_model(
            extraction,
            classify_protected_from_domain,
            classify_blocked_from_domain,
        ):
            continue
        custom_id = f"site_{seq:06d}"
        mapping[custom_id] = extraction
        if extraction.source_quality == "protected":
            extraction.error = append_error(extraction.error, "classified_from_domain_only")
            logging.info(
                "Protected site will be classified from domain only: %s final_url=%s",
                extraction.original_url,
                extraction.homepage.final_url if extraction.homepage else "",
            )
            user_prompt = build_domain_only_prompt(
                extraction,
                min_categories=min_categories,
                max_categories=max_categories,
            )
        elif extraction.source_quality == "blocked":
            extraction.error = append_error(extraction.error, "classified_from_domain_only")
            logging.info(
                "Blocked site will be classified from domain only: %s final_url=%s",
                extraction.original_url,
                extraction.homepage.final_url if extraction.homepage else "",
            )
            user_prompt = build_domain_only_prompt(
                extraction,
                min_categories=min_categories,
                max_categories=max_categories,
            )
        else:
            user_prompt = build_user_prompt(
                extraction,
                max_prompt_chars=max_prompt_chars,
                min_categories=min_categories,
                max_categories=max_categories,
            )
        prompt_chars = len(system_prompt) + len(user_prompt)
        extraction.prompt_chars = prompt_chars
        extraction.estimated_input_tokens = estimate_tokens_from_text(system_prompt + "\n" + user_prompt)
        extraction.token_count_method = "estimated"
        if count_tokens and client is not None:
            exact_tokens, method = count_input_tokens_exact(client, model, system_prompt, user_prompt)
            if exact_tokens > 0:
                extraction.exact_input_tokens = exact_tokens
                extraction.token_count_method = method
            else:
                extraction.token_count_method = method
        requests_payload.append(
            Request(
                custom_id=custom_id,
                params=MessageCreateParamsNonStreaming(
                    model=model,
                    max_tokens=max_tokens,
                    temperature=0,
                    system=system_prompt,
                    messages=[{"role": "user", "content": user_prompt}],
                ),
            )
        )
    return requests_payload, mapping


def submit_batch(client: Any, requests_payload: List[Any]) -> Any:
    if not requests_payload:
        return None
    return client.messages.batches.create(requests=requests_payload)


def poll_batch(client: Any, batch_id: str, poll_interval: int) -> Any:
    while True:
        batch = client.messages.batches.retrieve(batch_id)
        logging.info("Batch %s status=%s counts=%s", batch_id, batch.processing_status, getattr(batch, "request_counts", None))
        if batch.processing_status == "ended":
            return batch
        time.sleep(poll_interval)


def extract_text_from_message(message: Any) -> str:
    parts: List[str] = []
    for block in getattr(message, "content", []) or []:
        if getattr(block, "type", None) == "text":
            parts.append(getattr(block, "text", ""))
        elif isinstance(block, dict) and block.get("type") == "text":
            parts.append(block.get("text", ""))
    return "\n".join(parts).strip()


def read_batch_results(
    client: Any,
    batch_id: str,
    batch_status: str,
    custom_id_map: Dict[str, SiteExtraction],
    model_name: str,
    max_categories: int = DEFAULT_MAX_CATEGORIES,
) -> Dict[int, ClassificationResult]:
    by_row: Dict[int, ClassificationResult] = {}
    for item in client.messages.batches.results(batch_id):
        custom_id = item.custom_id
        extraction = custom_id_map.get(custom_id)
        if not extraction:
            logging.warning("Result custom_id not found in mapping: %s", custom_id)
            continue

        result_type = getattr(item.result, "type", "")
        if result_type == "succeeded":
            message = item.result.message
            raw_text = extract_text_from_message(message)
            parsed = parse_and_validate_model_result(
                row_index=extraction.row_index,
                original_url=extraction.original_url,
                custom_id=custom_id,
                model_name=model_name,
                batch_id=batch_id,
                batch_status=batch_status,
                raw_text=raw_text,
                max_categories=min(5, max_categories)
                if extraction.source_quality in {"protected", "blocked"}
                else max_categories,
            )
            usage = getattr(message, "usage", None)
            parsed.input_tokens = int(getattr(usage, "input_tokens", 0) or 0)
            parsed.output_tokens = int(getattr(usage, "output_tokens", 0) or 0)
            parsed.estimated_cost_usd = calculate_batch_cost_usd(model_name, parsed.input_tokens, parsed.output_tokens)
            by_row[extraction.row_index] = parsed
        else:
            err = ""
            try:
                err = json.dumps(item.result.model_dump(), ensure_ascii=False)
            except Exception:
                err = str(item.result)
            by_row[extraction.row_index] = ClassificationResult(
                row_index=extraction.row_index,
                original_url=extraction.original_url,
                custom_id=custom_id,
                niche=["UNKNOWN"],
                categories=[],
                language=LANGUAGE_UNKNOWN,
                model_name=model_name,
                batch_id=batch_id,
                batch_status=batch_status,
                model_error=f"batch_result_{result_type}:{err[:500]}",
            )
    return by_row


# ===================== EXCEL OUTPUT =====================
DEFAULT_RESULT_COLUMNS = [
    "SourceRowNumber",
    "Niche",
    "Categories",
    "Language",
    "SourceQuality",
    "FinalUrl",
    "Error",
]

DEBUG_RESULT_COLUMNS = [
    "HomepageFetchMethod",
    "Redirected",
    "HttpStatus",
    "HomepageTextLength",
    "InternalPagesRequested",
    "InternalPagesHttpSucceeded",
    "InternalPagesSeleniumTried",
    "InternalPagesSeleniumSucceeded",
    "InternalPagesSucceeded",
    "ChromeVersionMainUsed",
    "TotalExtractedTextLength",
    "ModelName",
    "BatchId",
    "BatchStatus",
    "ModelError",
    "PromptChars",
    "EstimatedInputTokens",
    "ExactInputTokens",
    "ActualInputTokens",
    "OutputTokens",
    "EstimatedCostUsd",
    "TokenCountMethod",
    "ProcessingTimeMs",
    "RawNiche",
    "RawCategories",
    "NormalizationWarnings",
    "RawModelText",
]

COLUMN_WIDTHS = {
    "SourceRowNumber": 16,
    "Niche": 28,
    "Categories": 60,
    "Language": 14,
    "SourceQuality": 16,
    "FinalUrl": 50,
    "Error": 50,
    "HomepageFetchMethod": 22,
    "Redirected": 14,
    "HttpStatus": 14,
    "HomepageTextLength": 22,
    "InternalPagesRequested": 24,
    "InternalPagesHttpSucceeded": 28,
    "InternalPagesSeleniumTried": 28,
    "InternalPagesSeleniumSucceeded": 32,
    "InternalPagesSucceeded": 24,
    "ChromeVersionMainUsed": 24,
    "TotalExtractedTextLength": 26,
    "ModelName": 24,
    "BatchId": 34,
    "BatchStatus": 18,
    "ModelError": 50,
    "PromptChars": 16,
    "EstimatedInputTokens": 24,
    "ExactInputTokens": 20,
    "ActualInputTokens": 20,
    "OutputTokens": 16,
    "EstimatedCostUsd": 20,
    "TokenCountMethod": 20,
    "ProcessingTimeMs": 20,
    "RawNiche": 40,
    "RawCategories": 60,
    "NormalizationWarnings": 60,
    "RawModelText": 60,
}


def get_result_columns(debug_output: bool) -> List[str]:
    if debug_output:
        return DEFAULT_RESULT_COLUMNS + DEBUG_RESULT_COLUMNS
    return DEFAULT_RESULT_COLUMNS


def detect_url_column(df: pd.DataFrame, explicit: Optional[str]) -> str:
    if explicit:
        if explicit not in df.columns:
            raise RuntimeError(f"URL column '{explicit}' not found. Available columns: {list(df.columns)}")
        return explicit
    for col in ["Website", "URL", "Url", "Domain", "domain"]:
        if col in df.columns:
            return col
    raise RuntimeError("Input file must contain one of these columns: Website, URL, Url, Domain")


def empty_classification_for_extraction(extraction: SiteExtraction, model_name: str, batch_id: str, batch_status: str) -> ClassificationResult:
    model_error = ""
    if extraction.source_quality == "failed":
        model_error = "not_sent_to_model_due_to_failed_extraction"
    elif extraction.source_quality == "protected":
        model_error = "not_sent_to_model_due_to_bot_protection"
    elif extraction.source_quality == "blocked":
        model_error = "not_sent_to_model_due_to_access_block"
    return ClassificationResult(
        row_index=extraction.row_index,
        original_url=extraction.original_url,
        niche=["UNKNOWN"],
        categories=[],
        language=LANGUAGE_UNKNOWN,
        model_name=model_name,
        batch_id=batch_id,
        batch_status=batch_status,
        model_error=model_error,
    )


def validate_output_path(output_path: str) -> None:
    if os.path.splitext(output_path)[1].lower() != ".xlsx":
        raise RuntimeError("Output file must be an .xlsx file. Example: --output sites_with_categories_v2.xlsx")


def excel_value(value: Any) -> Any:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, (list, tuple, set)):
        return ", ".join(str(v) for v in value if v is not None and str(v).strip())
    return value


def excel_text(value: Any) -> str:
    safe = excel_value(value)
    if safe is None:
        return ""
    return str(safe)


def excel_long_text(value: Any, max_len: int = 5000) -> str:
    text = excel_text(value)
    if len(text) > max_len:
        return text[: max_len - 1] + "…"
    return text


def excel_bool(value: Optional[bool]) -> str:
    if value is None:
        return ""
    return "TRUE" if bool(value) else "FALSE"


def excel_int(value: Optional[int]) -> Any:
    safe = excel_value(value)
    if safe == "":
        return ""
    return int(safe)


def excel_float(value: Optional[float], decimals: int = 8) -> Any:
    safe = excel_value(value)
    if safe == "":
        return ""
    return round(float(safe), decimals)


def build_result_values(
    extraction: SiteExtraction,
    classification: ClassificationResult,
    model_name: str,
    batch_id: str,
    batch_status: str,
    extract_only: bool,
    chrome_version_main: Optional[int] = None,
) -> Dict[str, Any]:
    homepage = extraction.homepage
    detected_language = homepage.html_lang if homepage and homepage.html_lang else LANGUAGE_UNKNOWN
    language = classification.language or detected_language or LANGUAGE_UNKNOWN
    if language == LANGUAGE_UNKNOWN and detected_language:
        language = detected_language

    errors = [e for e in [extraction.error] if e]
    if extraction.source_quality not in {"protected", "blocked"} and classification.model_error:
        errors.append(classification.model_error)
    return {
        "SourceRowNumber": excel_int(extraction.row_index + 1),
        "Niche": "" if extract_only else excel_text(classification.niche or ["UNKNOWN"]),
        "Categories": excel_text(classification.categories or []),
        "Language": excel_text(language or LANGUAGE_UNKNOWN),
        "Error": excel_text("; ".join(errors)),
        "HomepageFetchMethod": excel_text(homepage.fetch_method if homepage else "failed"),
        "FinalUrl": excel_text(homepage.final_url if homepage else ""),
        "Redirected": excel_bool(homepage.redirected if homepage else None),
        "HttpStatus": excel_int(homepage.http_status if homepage and homepage.http_status is not None else None),
        "HomepageTextLength": excel_int(homepage.text_length if homepage else 0),
        "InternalPagesRequested": excel_int(extraction.internal_pages_requested),
        "InternalPagesHttpSucceeded": excel_int(extraction.internal_pages_http_succeeded),
        "InternalPagesSeleniumTried": excel_int(extraction.internal_pages_selenium_tried),
        "InternalPagesSeleniumSucceeded": excel_int(extraction.internal_pages_selenium_succeeded),
        "InternalPagesSucceeded": excel_int(extraction.internal_pages_succeeded),
        "ChromeVersionMainUsed": excel_int(chrome_version_main),
        "TotalExtractedTextLength": excel_int(extraction.total_text_length),
        "SourceQuality": excel_text(extraction.source_quality),
        "ModelName": "" if extract_only else excel_text(classification.model_name or model_name),
        "BatchId": "" if extract_only else excel_text(classification.batch_id or batch_id),
        "BatchStatus": excel_text(classification.batch_status or batch_status),
        "ModelError": excel_text(classification.model_error),
        "PromptChars": excel_int(extraction.prompt_chars),
        "EstimatedInputTokens": excel_int(extraction.estimated_input_tokens),
        "ExactInputTokens": excel_int(extraction.exact_input_tokens),
        "ActualInputTokens": "" if extract_only else excel_int(classification.input_tokens),
        "OutputTokens": "" if extract_only else excel_int(classification.output_tokens),
        "EstimatedCostUsd": "" if extract_only else excel_float(classification.estimated_cost_usd, 8),
        "TokenCountMethod": excel_text(extraction.token_count_method),
        "ProcessingTimeMs": excel_int(extraction.processing_time_ms),
        "RawNiche": "" if extract_only else excel_text(classification.raw_niche or []),
        "RawCategories": "" if extract_only else excel_text(classification.raw_categories or []),
        "NormalizationWarnings": "" if extract_only else excel_text("; ".join(classification.normalization_warnings)),
        "RawModelText": "" if extract_only else excel_long_text(classification.raw_model_text, 5000),
    }


def build_output_dataframe(
    df: pd.DataFrame,
    extractions: List[SiteExtraction],
    classifications_by_row: Dict[int, ClassificationResult],
    model_name: str,
    batch_id: str,
    batch_status: str,
    output_scope: str = "processed",
    debug_output: bool = False,
    chrome_version_main: Optional[int] = None,
) -> pd.DataFrame:
    original_columns = list(df.columns)
    result_columns = get_result_columns(debug_output)
    fieldnames = original_columns + [col for col in result_columns if col not in original_columns]
    if len(fieldnames) != len(set(fieldnames)):
        raise RuntimeError("Output DataFrame has duplicate columns")

    extractions_by_row = {extraction.row_index: extraction for extraction in extractions}
    if output_scope == "processed":
        row_indexes = sorted(extractions_by_row.keys())
    elif output_scope == "all":
        row_indexes = list(df.index)
    else:
        raise RuntimeError(f"Unsupported output_scope: {output_scope}")

    logging.info(
        "Building output dataframe: scope=%s debug_output=%s processed_rows=%s output_rows=%s",
        output_scope,
        debug_output,
        len(extractions_by_row),
        len(row_indexes),
    )

    rows: List[Dict[str, Any]] = []

    for input_index in row_indexes:
        input_row = df.loc[input_index]
        output_row: Dict[str, Any] = {}
        for col in original_columns:
            output_row[col] = excel_value(input_row.get(col, ""))

        extraction = extractions_by_row.get(input_index)
        if extraction:
            extract_only = batch_status == "extract_only" and input_index not in classifications_by_row
            classification = classifications_by_row.get(input_index)
            if classification is None and extract_only:
                homepage = extraction.homepage
                classification = ClassificationResult(
                    row_index=extraction.row_index,
                    original_url=extraction.original_url,
                    niche=[],
                    categories=[],
                    language=homepage.html_lang if homepage and homepage.html_lang else LANGUAGE_UNKNOWN,
                    batch_status="extract_only",
                )
            elif classification is None:
                classification = empty_classification_for_extraction(
                    extraction, model_name=model_name, batch_id=batch_id, batch_status=batch_status
                )
            output_row.update(
                build_result_values(
                    extraction=extraction,
                    classification=classification,
                    model_name=model_name,
                    batch_id=batch_id,
                    batch_status=batch_status,
                    extract_only=extract_only,
                    chrome_version_main=chrome_version_main,
                )
            )
        else:
            for col in result_columns:
                output_row[col] = excel_int(input_index + 1) if col == "SourceRowNumber" else ""

        rows.append(output_row)

    output_df = pd.DataFrame(rows, columns=fieldnames)
    if len(output_df) != len(row_indexes):
        raise RuntimeError(f"Output row count mismatch: expected {len(row_indexes)}, got {len(output_df)}")
    if len(output_df.columns) != len(set(output_df.columns)):
        raise RuntimeError("Output DataFrame has duplicate columns")
    return output_df


def write_output_excel(output_path: str, output_df: pd.DataFrame, enable_filter: bool = False) -> None:
    if len(output_df.columns) != len(set(output_df.columns)):
        raise RuntimeError("Output DataFrame has duplicate columns")
    output_df.to_excel(output_path, index=False, engine="openpyxl")

    from openpyxl import load_workbook

    wb = load_workbook(output_path)
    ws = wb.active
    ws.title = "Results"

    apply_excel_formatting(ws, enable_filter=enable_filter)

    wb.save(output_path)
    logging.info("Formatted Excel output for review: freeze_panes=A2")
    if enable_filter:
        logging.info("Excel autofilter enabled.")
    logging.info("Saved Excel output to %s rows=%s", output_path, len(output_df))


def apply_excel_formatting(ws: Any, enable_filter: bool = False) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    min_width = 12
    max_width = 50
    sample_limit = min(ws.max_row, 101)

    ws.freeze_panes = "A2"
    ws.sheet_view.zoomScale = 90
    if enable_filter:
        ws.auto_filter.ref = ws.dimensions

    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = header_fill
    ws.row_dimensions[1].height = 36

    headers_by_index = {cell.column: str(cell.value or "") for cell in ws[1]}
    for col_idx in range(1, ws.max_column + 1):
        header = headers_by_index.get(col_idx, "")
        col_letter = get_column_letter(col_idx)

        if header in COLUMN_WIDTHS:
            width = COLUMN_WIDTHS[header]
        else:
            sampled_lengths = [len(header)]
            for row_idx in range(2, sample_limit + 1):
                value = ws.cell(row=row_idx, column=col_idx).value
                if value is None:
                    continue
                sampled_lengths.append(max(len(line) for line in str(value).splitlines() or [""]))
            width = min(max(max(sampled_lengths) + 2, min_width), max_width)

        ws.column_dimensions[col_letter].width = width

    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 18
        for cell in ws[row_idx]:
            cell.alignment = Alignment(vertical="top", wrap_text=False)


# ===================== MAIN FLOW =====================
def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Classify website Niche/Categories/Language with Claude Batch API.")
    parser.add_argument("--input", default=DEFAULT_INPUT, help="Input Excel file path.")
    parser.add_argument("--output", default=DEFAULT_OUTPUT, help="Output Excel file path.")
    parser.add_argument("--url-column", default=DEFAULT_URL_COLUMN, help="Optional URL column name. Auto-detects Website/URL/Domain by default.")
    parser.add_argument("--checkpoint", default=DEFAULT_CHECKPOINT, help="Checkpoint JSON path.")
    parser.add_argument("--resume", action="store_true", help="Resume from checkpoint if it exists.")
    parser.add_argument("--max-sites", type=int, default=DEFAULT_MAX_SITES, help="Max rows to process. Default 20 to avoid accidental large paid runs.")
    parser.add_argument("--start-row", type=int, default=DEFAULT_START_ROW, help="Zero-based row offset in the Excel dataframe.")
    parser.add_argument(
        "--output-scope",
        choices=["processed", "all"],
        default="processed",
        help="Controls which input rows are written to output. processed=only rows actually processed by this run/checkpoint, all=all input rows with blank result columns for unprocessed rows.",
    )
    parser.add_argument("--concurrency", type=int, default=DEFAULT_CONCURRENCY, help="Concurrent website extraction workers.")
    parser.add_argument("--request-timeout", type=int, default=DEFAULT_REQUEST_TIMEOUT, help="HTTP request timeout seconds.")
    parser.add_argument("--internal-pages", type=int, default=DEFAULT_INTERNAL_PAGES, help="Max internal pages to fetch per site. 0 disables light crawler.")
    parser.add_argument("--min-text-chars", type=int, default=DEFAULT_MIN_TEXT_CHARS, help="Minimum homepage extracted text chars before considering text usable.")
    parser.add_argument("--min-internal-text-chars", type=int, default=DEFAULT_MIN_INTERNAL_TEXT_CHARS, help="Minimum internal page extracted text chars before using it.")
    parser.add_argument("--no-selenium", action="store_true", help="Deprecated alias for --selenium-mode off.")
    parser.add_argument(
        "--selenium-mode",
        choices=["auto", "off", "required"],
        default="auto",
        help="Controls Selenium fallback. auto=use if available, off=HTTP only, required=fail fast if Selenium is unavailable.",
    )
    parser.add_argument("--selenium-timeout", type=int, default=DEFAULT_SELENIUM_TIMEOUT, help="Selenium page load / wait timeout seconds.")
    parser.add_argument("--selenium-hard-timeout", type=int, default=DEFAULT_SELENIUM_HARD_TIMEOUT, help="Hard timeout for one Selenium fetch.")
    parser.add_argument("--selenium-internal-fallback", action="store_true", help="Allow Selenium fallback for weak/failed internal pages.")
    parser.add_argument("--selenium-internal-limit", type=int, default=DEFAULT_SELENIUM_INTERNAL_LIMIT, help="Max internal pages per site that may use Selenium if fallback enabled.")
    parser.add_argument(
        "--chrome-version-main",
        type=int,
        default=None,
        help="Chrome major version passed to undetected_chromedriver, e.g. 147.",
    )
    parser.add_argument("--model", choices=MODEL_CHOICES, default=DEFAULT_MODEL, help="Claude model. Default claude-haiku-4-5.")
    parser.add_argument("--api-key", default=ANTHROPIC_API_KEY, help="Anthropic API key. Overrides ANTHROPIC_API_KEY constant and env var.")
    parser.add_argument("--max-tokens", type=int, default=DEFAULT_MAX_TOKENS, help="Max output tokens per site classification request.")
    parser.add_argument(
        "--min-categories",
        type=int,
        default=DEFAULT_MIN_CATEGORIES,
        help="Minimum number of English Categories/search tags requested from Claude when enough content is available.",
    )
    parser.add_argument(
        "--max-categories",
        type=int,
        default=DEFAULT_MAX_CATEGORIES,
        help="Maximum number of English Categories/search tags kept after validation.",
    )
    parser.add_argument("--max-prompt-chars", type=int, default=DEFAULT_MAX_PROMPT_CHARS, help="Max user prompt characters per site before dropping internal pages/trimming text.")
    parser.add_argument("--count-tokens", action="store_true", help="Call Anthropic token counting API before batch creation and write ExactInputTokens.")
    parser.add_argument(
        "--classify-protected-from-domain",
        action="store_true",
        help="For bot-protected sites, ask Claude to infer only a cautious niche/categories from domain/final URL. SourceQuality remains protected.",
    )
    parser.add_argument(
        "--classify-blocked-from-domain",
        action="store_true",
        help="For access-blocked sites, ask Claude to infer only a cautious niche/categories from domain/final URL. SourceQuality remains blocked.",
    )
    parser.add_argument("--poll-interval", type=int, default=DEFAULT_POLL_INTERVAL, help="Batch polling interval seconds.")
    parser.add_argument("--submit-only", action="store_true", help="Create Anthropic batch and exit without polling/results.")
    parser.add_argument("--extract-only", action="store_true", help="Only fetch/crawl/extract and write analytics; do not call Anthropic.")
    parser.add_argument(
        "--debug-output",
        action="store_true",
        help="Include technical/debug columns in the output Excel.",
    )
    parser.add_argument(
        "--excel-filter",
        action="store_true",
        help="Enable Excel autofilter on the output worksheet.",
    )
    parser.add_argument("--log-file", default=DEFAULT_LOG_FILE, help="Log file path.")
    parser.add_argument("--proxy", action="append", default=DEFAULT_PROXY, help="Proxy URL. Can be passed multiple times.")
    args = parser.parse_args()
    if args.min_categories < 1:
        parser.error("--min-categories must be at least 1")
    if args.max_categories < args.min_categories:
        parser.error("--max-categories must be greater than or equal to --min-categories")
    if args.max_categories > 40:
        parser.error("--max-categories must be <= 40 to keep prompts/results useful and costs controlled")
    if args.no_selenium:
        if args.selenium_mode == "required":
            parser.error("--no-selenium conflicts with --selenium-mode required")
        args.selenium_mode = "off"
    args.internal_pages = max(0, min(10, args.internal_pages))
    args.selenium_internal_limit = max(0, min(args.internal_pages, args.selenium_internal_limit))
    args.proxies = args.proxy or []
    return args


def serialize_extractions(extractions: List[SiteExtraction]) -> List[Dict[str, Any]]:
    return [asdict(x) for x in extractions]


def deserialize_page(data: Optional[Dict[str, Any]]) -> Optional[PageExtract]:
    return PageExtract(**data) if data else None


def deserialize_extractions(items: List[Dict[str, Any]]) -> List[SiteExtraction]:
    result = []
    for item in items:
        item = dict(item)
        item["homepage"] = deserialize_page(item.get("homepage"))
        item["internal_pages"] = [PageExtract(**p) for p in item.get("internal_pages", [])]
        result.append(SiteExtraction(**item))
    return result


def run_extraction(
    df: pd.DataFrame,
    url_col: str,
    args: argparse.Namespace,
    chrome_version_main: Optional[int] = None,
) -> List[SiteExtraction]:
    selected = df.iloc[args.start_row : args.start_row + args.max_sites]
    tasks = [(idx, row[url_col]) for idx, row in selected.iterrows()]
    extractions: List[SiteExtraction] = []

    logging.info("Starting extraction: rows=%s concurrency=%s internal_pages=%s", len(tasks), args.concurrency, args.internal_pages)
    with ThreadPoolExecutor(max_workers=max(1, args.concurrency)) as executor:
        futures = {executor.submit(crawl_site, idx, url, args, chrome_version_main): idx for idx, url in tasks}
        for future in tqdm(as_completed(futures), total=len(futures), desc="Extracting"):
            idx = futures[future]
            try:
                extraction = future.result()
            except Exception as e:
                logging.error("Extraction failed for row %s: %s\n%s", idx, e, traceback.format_exc())
                extraction = SiteExtraction(
                    row_index=idx,
                    original_url=str(df.at[idx, url_col] if url_col in df.columns else ""),
                    normalized_start_url=normalize_input_url(df.at[idx, url_col] if url_col in df.columns else ""),
                    error=f"future_exception:{type(e).__name__}:{e}",
                    source_quality="failed",
                )
            extractions.append(extraction)

    extractions.sort(key=lambda x: x.row_index)
    return extractions


def log_usage_summary(classifications_by_row: Dict[int, ClassificationResult], model: str) -> None:
    input_tokens = sum(x.input_tokens for x in classifications_by_row.values())
    output_tokens = sum(x.output_tokens for x in classifications_by_row.values())
    estimated_cost = calculate_batch_cost_usd(model, input_tokens, output_tokens)
    logging.info(
        "Claude usage summary: actual_input_tokens=%s output_tokens=%s estimated_batch_cost_usd=%.6f",
        input_tokens, output_tokens, estimated_cost,
    )


def main() -> None:
    args = parse_args()
    setup_logging(args.log_file)

    logging.info("parser_v2 starting. model=%s max_sites=%s", args.model, args.max_sites)
    chrome_version_main = resolve_chrome_version_main(args.chrome_version_main)
    selenium_availability = check_selenium_availability(args.selenium_mode)
    args.selenium_availability = selenium_availability
    if selenium_availability.user_message:
        if selenium_availability.available or args.selenium_mode == "off":
            logging.info(selenium_availability.user_message)
        else:
            logging.warning(selenium_availability.user_message)
    validate_output_path(args.output)
    if not os.path.exists(args.input):
        raise RuntimeError(f"Input file not found: {args.input}")

    df = pd.read_excel(args.input)
    url_col = detect_url_column(df, args.url_column)

    checkpoint = load_checkpoint(args.checkpoint) if args.resume else {}
    batch_id = checkpoint.get("batch_id", "")
    batch_status = checkpoint.get("batch_status", "")
    custom_id_to_row = checkpoint.get("custom_id_to_row", {})

    if checkpoint.get("extractions"):
        logging.info("Loaded extractions from checkpoint: %s", len(checkpoint["extractions"]))
        extractions = deserialize_extractions(checkpoint["extractions"])
    else:
        extractions = run_extraction(df, url_col, args, chrome_version_main=chrome_version_main)
        save_checkpoint(
            args.checkpoint,
            {
                "version": 2,
                "model": args.model,
                "input": args.input,
                "output": args.output,
                "url_column": url_col,
                "extractions": serialize_extractions(extractions),
                "batch_id": "",
                "batch_status": "not_created",
                "custom_id_to_row": {},
            },
        )

    protected_count = len([x for x in extractions if x.source_quality == "protected"])
    blocked_count = len([x for x in extractions if x.source_quality == "blocked"])
    logging.info("Protected pages detected: %s", protected_count)
    logging.info("Blocked/access-denied pages detected: %s", blocked_count)
    if args.classify_protected_from_domain:
        logging.info("Protected domain-only classification is enabled.")
    else:
        logging.info("Protected pages will not be sent to Claude.")
    if args.classify_blocked_from_domain:
        logging.info("Blocked domain-only classification is enabled.")
    else:
        logging.info("Blocked pages will not be sent to Claude.")

    if args.extract_only:
        output = build_output_dataframe(
            df,
            extractions,
            classifications_by_row={},
            model_name=args.model,
            batch_id="",
            batch_status="extract_only",
            output_scope=args.output_scope,
            debug_output=args.debug_output,
            chrome_version_main=chrome_version_main,
        )
        write_output_excel(args.output, output, enable_filter=args.excel_filter)
        return

    classifications_by_row: Dict[int, ClassificationResult] = {}

    # For rows that cannot be sent to Claude, write explicit fallback classifications.
    failed_extractions = [
        x
        for x in extractions
        if not should_send_extraction_to_model(
            x,
            args.classify_protected_from_domain,
            args.classify_blocked_from_domain,
        )
    ]
    for extraction in failed_extractions:
        classifications_by_row[extraction.row_index] = empty_classification_for_extraction(
            extraction, model_name=args.model, batch_id=batch_id, batch_status=batch_status or "not_sent"
        )

    if not batch_id:
        client = create_anthropic_client(args.api_key)
        requests_payload, custom_id_map = create_batch_requests(
            extractions,
            model=args.model,
            max_tokens=args.max_tokens,
            max_prompt_chars=args.max_prompt_chars,
            min_categories=args.min_categories,
            max_categories=args.max_categories,
            client=client,
            count_tokens=args.count_tokens,
            classify_protected_from_domain=args.classify_protected_from_domain,
            classify_blocked_from_domain=args.classify_blocked_from_domain,
        )
        estimated_input_tokens = sum(x.estimated_input_tokens for x in extractions)
        exact_input_tokens = sum(x.exact_input_tokens for x in extractions)
        input_for_estimate = exact_input_tokens or estimated_input_tokens
        estimated_input_cost = calculate_batch_cost_usd(args.model, input_for_estimate, 0)
        logging.info(
            "Prepared Claude batch requests=%s estimated_input_tokens=%s exact_input_tokens=%s estimated_input_cost_usd=%.6f",
            len(requests_payload), estimated_input_tokens, exact_input_tokens, estimated_input_cost,
        )

        if not requests_payload:
            logging.warning("No valid extracted sites to send to Claude.")
            output = build_output_dataframe(
                df,
                extractions,
                classifications_by_row=classifications_by_row,
                model_name=args.model,
                batch_id="",
                batch_status="no_valid_requests",
                output_scope=args.output_scope,
                debug_output=args.debug_output,
                chrome_version_main=chrome_version_main,
            )
            write_output_excel(args.output, output, enable_filter=args.excel_filter)
            return

        logging.info("Creating Anthropic batch with %s requests...", len(requests_payload))
        batch = submit_batch(client, requests_payload)
        batch_id = batch.id
        batch_status = batch.processing_status
        custom_id_to_row = {custom_id: extraction.row_index for custom_id, extraction in custom_id_map.items()}
        save_checkpoint(
            args.checkpoint,
            {
                "version": 2,
                "model": args.model,
                "input": args.input,
                "output": args.output,
                "url_column": url_col,
                "extractions": serialize_extractions(extractions),
                "batch_id": batch_id,
                "batch_status": batch_status,
                "custom_id_to_row": custom_id_to_row,
            },
        )
        logging.info("Batch created: %s status=%s", batch_id, batch_status)
        if args.submit_only:
            output = build_output_dataframe(
                df,
                extractions,
                classifications_by_row=classifications_by_row,
                model_name=args.model,
                batch_id=batch_id,
                batch_status=batch_status,
                output_scope=args.output_scope,
                debug_output=args.debug_output,
                chrome_version_main=chrome_version_main,
            )
            write_output_excel(args.output, output, enable_filter=args.excel_filter)
            return
    else:
        client = create_anthropic_client(args.api_key)
        # Recreate custom_id_map from checkpoint row mapping.
        extraction_by_row = {x.row_index: x for x in extractions}
        custom_id_map = {
            custom_id: extraction_by_row[row]
            for custom_id, row in custom_id_to_row.items()
            if row in extraction_by_row
        }
        logging.info("Resuming batch: %s", batch_id)

    ended_batch = poll_batch(client, batch_id, poll_interval=args.poll_interval)
    batch_status = ended_batch.processing_status
    save_checkpoint(
        args.checkpoint,
        {
            "version": 2,
            "model": args.model,
            "input": args.input,
            "output": args.output,
            "url_column": url_col,
            "extractions": serialize_extractions(extractions),
            "batch_id": batch_id,
            "batch_status": batch_status,
            "custom_id_to_row": custom_id_to_row,
        },
    )

    model_results = read_batch_results(
        client,
        batch_id=batch_id,
        batch_status=batch_status,
        custom_id_map=custom_id_map,
        model_name=args.model,
        max_categories=args.max_categories,
    )
    classifications_by_row.update(model_results)
    log_usage_summary(classifications_by_row, args.model)

    output = build_output_dataframe(
        df,
        extractions,
        classifications_by_row=classifications_by_row,
        model_name=args.model,
        batch_id=batch_id,
        batch_status=batch_status,
        output_scope=args.output_scope,
        debug_output=args.debug_output,
        chrome_version_main=chrome_version_main,
    )
    write_output_excel(args.output, output, enable_filter=args.excel_filter)
    logging.info("Done.")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        logging.warning("Interrupted by user.")
        raise
    except Exception as e:
        logging.error("Fatal error: %s\n%s", e, traceback.format_exc())
        raise
