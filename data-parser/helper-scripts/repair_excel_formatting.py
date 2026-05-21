import os
import shutil
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


COLUMN_WIDTHS = {
    "SourceRowNumber": 16,
    "Niche": 28,
    "Categories": 60,
    "Language": 14,
    "SourceQuality": 16,
    "FinalUrl": 50,
    "Error": 50,
    "HomepageFetchMethod": 22,
    "Redirected": 14,
    "HttpStatus": 14,
    "HomepageTextLength": 22,
    "InternalPagesRequested": 24,
    "InternalPagesHttpSucceeded": 28,
    "InternalPagesSeleniumTried": 28,
    "InternalPagesSeleniumSucceeded": 32,
    "InternalPagesSucceeded": 24,
    "ChromeVersionMainUsed": 24,
    "TotalExtractedTextLength": 26,
    "ModelName": 24,
    "BatchId": 34,
    "BatchStatus": 18,
    "ModelError": 50,
    "PromptChars": 16,
    "EstimatedInputTokens": 24,
    "ExactInputTokens": 20,
    "ActualInputTokens": 20,
    "OutputTokens": 16,
    "EstimatedCostUsd": 20,
    "TokenCountMethod": 20,
    "ProcessingTimeMs": 20,
    "RawNiche": 40,
    "RawCategories": 60,
    "NormalizationWarnings": 60,
    "RawModelText": 60,
}


def apply_excel_formatting(path: str) -> None:
    source = Path(path)
    if not source.exists():
        raise FileNotFoundError(f"File not found: {source}")

    backup = source.with_suffix(".backup.xlsx")
    if not backup.exists():
        shutil.copy2(source, backup)
        print(f"Backup created: {backup}")

    print(f"Loading workbook: {source}")
    wb = load_workbook(source)
    ws = wb.active
    ws.title = "Results"

    ws.freeze_panes = "A2"
    ws.sheet_view.zoomScale = 90

    if ws.max_row >= 1 and ws.max_column >= 1:
        last_col = get_column_letter(ws.max_column)
        ws.auto_filter.ref = f"A1:{last_col}{ws.max_row}"

    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = header_fill

    ws.row_dimensions[1].height = 36

    headers_by_index = {
        cell.column: str(cell.value or "")
        for cell in ws[1]
    }

    min_width = 12
    max_width = 50
    sample_limit = min(ws.max_row, 101)

    for col_idx in range(1, ws.max_column + 1):
        header = headers_by_index.get(col_idx, "")
        col_letter = get_column_letter(col_idx)

        if header in COLUMN_WIDTHS:
            width = COLUMN_WIDTHS[header]
        else:
            sampled_lengths = [len(header)]
            for row_idx in range(2, sample_limit + 1):
                value = ws.cell(row=row_idx, column=col_idx).value
                if value is None:
                    continue
                sampled_lengths.append(max(len(line) for line in str(value).splitlines() or [""]))
            width = min(max(max(sampled_lengths) + 2, min_width), max_width)

        ws.column_dimensions[col_letter].width = width

    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 18
        for cell in ws[row_idx]:
            cell.alignment = Alignment(vertical="top", wrap_text=False)

    print("Saving workbook...")
    wb.save(source)
    print(f"Done: {source}")


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage:")
        print("  py -3.11 repair_excel_formatting.py sites_cheap_v3_part_005.xlsx")
        sys.exit(1)

    apply_excel_formatting(sys.argv[1])
